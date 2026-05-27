import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colours ───────────────────────────────────────────────────
NAVY        = "1E3A5F"
AMBER       = "F59E0B"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F5F5F7"
LIGHT_BLUE  = "EFF6FF"
LIGHT_AMBER = "FFFBEB"
LIGHT_GREEN = "F0FDF4"
LIGHT_RED   = "FEF2F2"
LIGHT_PURP  = "F5F3FF"
MID_GRAY    = "E5E7EB"
TEXT_GRAY   = "6B7280"
GREEN       = "16A34A"
RED         = "DC2626"
PURPLE      = "7C3AED"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=None, size=11, italic=False):
    return Font(bold=bold, color=color or "000000", size=size, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border(style="thin", color=MID_GRAY):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def header_row(ws, row, cols, bg=NAVY, fg=WHITE, size=11):
    for col, val in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(bold=True, color=fg, size=size)
        c.alignment = align("center")
        c.border = border()

def data_row(ws, row, values, bg=WHITE, bold=False, color="000000", fmt=None):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(bold=bold, color=color)
        c.alignment = align("left", wrap=True)
        c.border = border()
        if fmt and col == fmt[0]:
            c.number_format = fmt[1]

def title_cell(ws, row, col, text, size=14, color=NAVY, bg=WHITE, bold=True, colspan=1):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(bold=bold, color=color, size=size)
    c.fill = fill(bg)
    c.alignment = align("left", "center")
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + colspan - 1)

def eur(ws, row, col, val=None, formula=None, bg=WHITE, bold=False, color="000000"):
    c = ws.cell(row=row, column=col, value=val if val is not None else formula)
    c.number_format = u'€#,##0.00'
    c.fill = fill(bg)
    c.font = font(bold=bold, color=color)
    c.alignment = align("right")
    c.border = border()

# ══════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ── SHEET 1: SUMMARY ──────────────────────────────────────────
ws1 = wb.active
ws1.title = "Summary"
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 40
ws1.row_dimensions[2].height = 20
ws1.row_dimensions[3].height = 20

# Title banner
ws1.merge_cells("A1:G1")
c = ws1["A1"]
c.value = "SyndicSage V5 — Cost & Systems Overview"
c.fill = fill(NAVY)
c.font = Font(bold=True, color=WHITE, size=18, name="Calibri")
c.alignment = Alignment(horizontal="center", vertical="center")

ws1.merge_cells("A2:G2")
c = ws1["A2"]
c.value = "All amounts in EUR · Updated 2026-05-27"
c.fill = fill(NAVY)
c.font = Font(color="B0C4DE", size=10, italic=True)
c.alignment = Alignment(horizontal="center", vertical="center")

ws1.row_dimensions[3].height = 12

# KPI boxes (row 4-8)
kpi_titles = ["One-Time Setup", "Monthly (MVP)", "Annual (MVP)", "Monthly at Scale", "Break-Even (€49/mo plan)"]
kpi_bgs    = [LIGHT_AMBER, LIGHT_BLUE, LIGHT_BLUE, LIGHT_RED, LIGHT_GREEN]
kpi_cols   = [1, 2, 3, 5, 6]

for i, (title, bg, col) in enumerate(zip(kpi_titles, kpi_bgs, kpi_cols)):
    ws1.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col)
    ws1.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col)
    ws1.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col)
    c = ws1.cell(row=4, column=col, value=title)
    c.fill = fill(bg); c.font = font(bold=True, color=NAVY, size=10)
    c.alignment = align("center")
    for r in range(4, 9):
        ws1.cell(row=r, column=col).fill = fill(bg)

# KPI values (linked to other sheets later with formulas)
kpi_vals = [
    ("='One-Time Costs'!B2", LIGHT_AMBER),
    ("='Monthly Costs'!C2", LIGHT_BLUE),
    ("='Monthly Costs'!C3", LIGHT_BLUE),
    ("='Monthly Costs'!C4", LIGHT_RED),
    ("='Break-Even'!B4", LIGHT_GREEN),
]
for (formula, bg), col in zip(kpi_vals, kpi_cols):
    c = ws1.cell(row=6, column=col, value=formula)
    c.number_format = u'€#,##0.00'
    c.fill = fill(bg)
    c.font = Font(bold=True, color=NAVY, size=16)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[6].height = 32

ws1.row_dimensions[8].height = 16

# Phase summary table
ws1.cell(row=9, column=1, value="Phase Cost Milestones").font = font(bold=True, color=NAVY, size=12)
ws1.row_dimensions[9].height = 22

header_row(ws1, 10, ["Phase", "Name", "New Services Added", "Extra Monthly Cost", "Cumulative Monthly"])
phase_data = [
    ("Phase 0-1", "Foundation + Shell",   "Supabase Pro, Railway, Vercel, GitHub Private",         "€61",  "€61"),
    ("Phase 2",   "Core Data",            "Google Workspace (already set up)",                      "€0",   "€61"),
    ("Phase 3",   "Communication",        "Resend (free tier), Sentry (free tier)",                 "€0",   "€61"),
    ("Phase 4",   "Intelligence",         "Anthropic API (usage-based), Upstash Redis",             "€30",  "€91"),
    ("Phase 5",   "Governance",           "No new services",                                        "€0",   "€91"),
    ("Phase 6",   "Portal",               "Resend scaling (more emails)",                           "€20",  "€111"),
    ("Phase 7",   "Mobile",               "Apple Developer ($99/yr), Expo EAS Build",               "€18",  "€129"),
    ("Phase 8",   "Launch",               "Vercel Pro (custom domain + bandwidth), DPA sign-offs",  "€20",  "€149"),
]
row_bgs = [LIGHT_BLUE, WHITE, LIGHT_BLUE, WHITE, LIGHT_BLUE, WHITE, LIGHT_BLUE, WHITE]
for i, (pd, bg) in enumerate(zip(phase_data, row_bgs)):
    r = 11 + i
    for col, val in enumerate(pd, 1):
        c = ws1.cell(row=r, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(size=10)
        c.alignment = align("left", wrap=True)
        c.border = border()
    ws1.row_dimensions[r].height = 22

# Column widths
for col, w in [(1,14),(2,18),(3,44),(4,20),(5,22),(6,22),(7,14)]:
    set_col_width(ws1, col, w)

# ── SHEET 2: MONTHLY COSTS ────────────────────────────────────
ws2 = wb.create_sheet("Monthly Costs")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:G1")
c = ws2["A1"]
c.value = "Monthly Recurring Costs"
c.fill = fill(NAVY); c.font = Font(bold=True, color=WHITE, size=14)
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 32

# Summary cells
summary_labels = [
    (2, "Total Monthly (MVP — Phase 0-3)",    "=SUMIF(E5:E100,\"MVP\",D5:D100)"),
    (3, "Total Monthly (Full — all phases)",  "=SUM(D5:D100)"),
    (4, "Total Monthly at Scale (estimate)",  "=SUMIF(E5:E100,\"Scale\",D5:D100)+SUMIF(E5:E100,\"MVP\",D5:D100)"),
]
for row, label, formula in summary_labels:
    c = ws2.cell(row=row, column=1, value=label)
    c.font = font(bold=True, color=NAVY, size=11)
    c.fill = fill(LIGHT_BLUE)
    c.border = border()
    c.alignment = align("left")
    eur(ws2, row, 2, formula=formula, bg=LIGHT_BLUE, bold=True, color=NAVY)
    eur(ws2, row, 3, formula=None, bg=LIGHT_BLUE)
    ws2.row_dimensions[row].height = 22

ws2.row_dimensions[4].height = 12

header_row(ws2, 5, ["Service", "Category", "Plan / Tier", "EUR / Month", "Phase", "Status", "Notes"])

monthly = [
    # Service, Category, Plan, EUR/mo, Phase, Status, Notes, bg
    ("Supabase",         "Infrastructure", "Pro (EU region)",           25,   "MVP",   "Required",  "PostgreSQL + Auth + Storage + Realtime. EU region mandatory for GDPR. Daily backups.", LIGHT_BLUE),
    ("Railway — API",    "Infrastructure", "Starter + usage",            8,   "MVP",   "Required",  "Hono API server. ~€8/mo for small workloads. Scales with usage.", WHITE),
    ("Railway — Worker", "Infrastructure", "Starter + usage",            7,   "MVP",   "Required",  "BullMQ worker (email, scan, PDF, AI jobs). Separate service from API.", LIGHT_BLUE),
    ("Vercel",           "Infrastructure", "Hobby (free → Pro at launch)",0,  "MVP",   "Free now",  "Hobby supports custom domains. Upgrade to Pro (€20/mo) at launch for bandwidth + team.", WHITE),
    ("GitHub",           "Development",   "Free → Team (private repo)",  0,   "MVP",   "Free now",  "Switch to private before launch. Team plan: $4/user/mo. Solo: free with GitHub Pro ($4/mo).", LIGHT_BLUE),
    ("Google Workspace", "Communication", "Business Starter",            6,   "MVP",   "Active",    "hello@syndicsage.com already set up. MX records propagated 2026-05-24.", WHITE),
    ("Upstash Redis",    "Infrastructure", "Free tier → Pay-as-you-go",  0,   "MVP",   "Free now",  "Rate limiting + BullMQ queue. Free: 10k commands/day. ~€5/mo at scale.", LIGHT_BLUE),
    ("Resend",           "Communication", "Free (3,000 emails/mo)",      0,   "MVP",   "Free now",  "Transactional email. Free tier sufficient until Portal launch. Scale plan: €20/mo for 50k.", WHITE),
    ("Sentry",           "Monitoring",    "Free (Developer)",            0,   "MVP",   "Free now",  "Error tracking + performance. Free tier: 5k errors/mo. Team: $26/mo when needed.", LIGHT_BLUE),
    ("Anthropic API",    "AI",            "Usage-based (Claude Sonnet)", 20,  "Phase 4","Required",  "AI Sage chat + extraction + summarization. €20/mo estimate for low usage. Scales with users.", WHITE),
    ("Expo EAS Build",   "Mobile",        "Free → Production",           0,   "Phase 7","Free now",  "Mobile build service. Free: 30 builds/mo. Production plan: $99/mo if needed.", LIGHT_BLUE),
    ("Vercel Pro",       "Infrastructure", "Pro",                        20,  "Phase 8","At launch", "Upgrade at launch for bandwidth, analytics, team access. €20/mo.", WHITE),
    ("GitHub Team",      "Development",   "Team (private)",              4,   "Phase 8","At launch", "Switch to private repo before launch. $4/user/mo.", LIGHT_BLUE),
    ("PostHog",          "Analytics",     "Free (1M events/mo)",         0,   "Optional","Optional", "Product analytics. Free tier covers early growth. No personal data needed.", WHITE),
    ("Plausible",        "Analytics",     "Starter",                     9,   "Optional","Optional", "Privacy-friendly analytics alternative to Google Analytics. GDPR-compliant.", LIGHT_BLUE),
]

for i, row_data in enumerate(monthly):
    r = 6 + i
    service, cat, plan, eur_val, phase, status, notes, bg = row_data
    values = [service, cat, plan, eur_val, phase, status, notes]
    for col, val in enumerate(values, 1):
        c = ws2.cell(row=r, column=col, value=val)
        c.fill = fill(bg)
        c.border = border()
        c.alignment = align("left", wrap=True)
        c.font = font(size=10)
    ws2.cell(row=r, column=4).number_format = u'€#,##0.00'
    ws2.cell(row=r, column=4).alignment = align("right")
    ws2.row_dimensions[r].height = 36

# Total row
total_row = 6 + len(monthly)
ws2.cell(row=total_row, column=1, value="TOTAL (all services)").font = font(bold=True, color=WHITE, size=11)
ws2.cell(row=total_row, column=1).fill = fill(NAVY)
ws2.cell(row=total_row, column=1).border = border()
for col in range(2, 8):
    ws2.cell(row=total_row, column=col).fill = fill(NAVY)
    ws2.cell(row=total_row, column=col).border = border()
c = ws2.cell(row=total_row, column=4, value=f"=SUM(D6:D{total_row-1})")
c.number_format = u'€#,##0.00'
c.fill = fill(NAVY)
c.font = font(bold=True, color=AMBER, size=12)
c.alignment = align("right")
c.border = border()
ws2.row_dimensions[total_row].height = 28

for col, w in [(1,22),(2,18),(3,26),(4,14),(5,12),(6,14),(7,54)]:
    set_col_width(ws2, col, w)

# ── SHEET 3: ONE-TIME COSTS ───────────────────────────────────
ws3 = wb.create_sheet("One-Time Costs")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:F1")
c = ws3["A1"]
c.value = "One-Time & Annual Costs"
c.fill = fill(NAVY); c.font = Font(bold=True, color=WHITE, size=14)
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 32

# Total summary
ws3.merge_cells("A2:C2")
c = ws3["A2"]
c.value = "Total One-Time (Pre-Launch Minimum)"
c.fill = fill(LIGHT_AMBER); c.font = font(bold=True, color=NAVY)
c.border = border(); c.alignment = align("left")
eur(ws3, 2, 4, formula="=SUMIF(E5:E100,\"One-time\",C5:C100)", bg=LIGHT_AMBER, bold=True, color=NAVY)
ws3.merge_cells("A3:C3")
c = ws3["A3"]
c.value = "Total Annual (recurring yearly)"
c.fill = fill(LIGHT_BLUE); c.font = font(bold=True, color=NAVY)
c.border = border(); c.alignment = align("left")
eur(ws3, 3, 4, formula="=SUMIF(E5:E100,\"Annual\",C5:C100)", bg=LIGHT_BLUE, bold=True, color=NAVY)
ws3.row_dimensions[2].height = 24; ws3.row_dimensions[3].height = 24
ws3.row_dimensions[4].height = 12

header_row(ws3, 5, ["Item", "Category", "EUR", "Priority", "Frequency", "Phase", "Notes"])

onetime = [
    # Item, Category, EUR, Priority, Frequency, Phase, Notes, bg
    ("Namecheap — syndicsage.com",      "Domain",        12,   "Required",  "Annual",   "Live",     "Already purchased. Renews ~€12/yr. Includes DNS management.", LIGHT_BLUE),
    ("Google Workspace",                "Communication", 72,   "Required",  "Annual",   "Active",   "€6/mo = €72/yr. hello@syndicsage.com. MX records set.", WHITE),
    ("EUIPO Trademark — SyndicSage",    "Legal",         900,  "High",      "One-time", "Pre-launch","Class 42+35. DIY via euipo.europa.eu Easy Filing. 4-6 months processing.", LIGHT_AMBER),
    ("Privacy Policy + Terms of Use",   "Legal",         0,    "Required",  "One-time", "Pre-launch","Use iubenda.com (€27/yr) or lawyer (€500-2000). Must be live before first user.", WHITE),
    ("GDPR DPA — Supabase",             "Legal/GDPR",    0,    "Required",  "One-time", "Pre-launch","Free. Sign via Supabase dashboard under Settings > Legal.", LIGHT_AMBER),
    ("GDPR DPA — Vercel",               "Legal/GDPR",    0,    "Required",  "One-time", "Pre-launch","Free. Sign via Vercel dashboard under Settings > Legal.", WHITE),
    ("GDPR DPA — Resend",               "Legal/GDPR",    0,    "Required",  "One-time", "Pre-launch","Free. Sign via Resend dashboard.", LIGHT_AMBER),
    ("GDPR DPA — Anthropic",            "Legal/GDPR",    0,    "Required",  "One-time", "Phase 4",  "Free. Anthropic provides standard DPA. Sign before AI features go live.", WHITE),
    ("GDPR DPA — Railway",              "Legal/GDPR",    0,    "Required",  "One-time", "Pre-launch","Free. Contact Railway support for DPA.", LIGHT_AMBER),
    ("Apple Developer Program",         "Mobile",        99,   "Required",  "Annual",   "Phase 7",  "Required to publish on App Store. $99/yr. Enables Apple Sign In on mobile.", WHITE),
    ("Google Play Console",             "Mobile",        25,   "Required",  "One-time", "Phase 7",  "One-time $25 registration fee to publish on Google Play.", LIGHT_AMBER),
    ("iubenda Privacy Policy",          "Legal",         27,   "Recommended","Annual",  "Pre-launch","GDPR-compliant privacy policy generator. Covers Belgium/EU. €27/yr.", WHITE),
    ("Sentry Team Plan",                "Monitoring",    312,  "Optional",  "Annual",   "Phase 3+", "$26/mo if free tier (5k errors/mo) is exceeded. Defer until needed.", LIGHT_AMBER),
    ("Semgrep Pro",                     "Security",      0,    "Optional",  "Annual",   "Deferred", "Free Community plan sufficient for solo founder. Pro when team grows.", WHITE),
    ("External Pentest",                "Security",      2000, "Deferred",  "One-time", "Phase 8+", "Trigger: first enterprise contract or €50k ARR. OWASP Top 10 + tenant isolation.", LIGHT_RED),
]

for i, row_data in enumerate(onetime):
    r = 6 + i
    item, cat, eur_val, priority, freq, phase, notes, bg = row_data
    values = [item, cat, eur_val, priority, freq, phase, notes]
    for col, val in enumerate(values, 1):
        c = ws3.cell(row=r, column=col, value=val)
        c.fill = fill(bg)
        c.border = border()
        c.alignment = align("left", wrap=True)
        c.font = font(size=10)
    ws3.cell(row=r, column=3).number_format = u'€#,##0.00'
    ws3.cell(row=r, column=3).alignment = align("right")
    # Color priority
    priority_colors = {"Required": GREEN, "High": AMBER, "Recommended": "2563EB", "Optional": TEXT_GRAY, "Deferred": RED}
    pcolor = priority_colors.get(priority, "000000")
    ws3.cell(row=r, column=4).font = font(bold=True, color=pcolor, size=10)
    ws3.row_dimensions[r].height = 36

total_row3 = 6 + len(onetime)
for col in range(1, 8):
    ws3.cell(row=total_row3, column=col).fill = fill(NAVY)
    ws3.cell(row=total_row3, column=col).border = border()
ws3.cell(row=total_row3, column=1, value="TOTAL ONE-TIME (Required only)").font = font(bold=True, color=WHITE)
ws3.cell(row=total_row3, column=1).fill = fill(NAVY)
c = ws3.cell(row=total_row3, column=3, value=f"=SUMIF(D6:D{total_row3-1},\"Required\",C6:C{total_row3-1})+SUMIF(D6:D{total_row3-1},\"High\",C6:C{total_row3-1})")
c.number_format = u'€#,##0.00'
c.fill = fill(NAVY)
c.font = font(bold=True, color=AMBER, size=12)
c.alignment = align("right")
c.border = border()
ws3.row_dimensions[total_row3].height = 28

for col, w in [(1,32),(2,18),(3,12),(4,14),(5,12),(6,14),(7,52)]:
    set_col_width(ws3, col, w)

# ── SHEET 4: ACCOUNTS & ACCESS ────────────────────────────────
ws4 = wb.create_sheet("Accounts & Access")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:G1")
c = ws4["A1"]
c.value = "All Accounts, Logins & Access — SyndicSage V5"
c.fill = fill(NAVY); c.font = Font(bold=True, color=WHITE, size=14)
c.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 32

header_row(ws4, 2, ["Service", "URL", "Account / Login", "Purpose", "Tier", "Status", "Monthly Cost"])

accounts = [
    # Service, URL, Account, Purpose, Tier, Status, Monthly Cost, bg
    ("Supabase",            "supabase.com",              "chrisndi-syn account",        "PostgreSQL DB + Auth + Storage + Realtime",   "Pro",              "Active",      "€25",   LIGHT_BLUE),
    ("Vercel",              "vercel.com",                "chrisndi-syn account",        "Frontend hosting — app.syndicsage.com",       "Hobby → Pro",      "Active",      "€0 → €20", WHITE),
    ("Railway",             "railway.app",               "chrisndi-syn account",        "Hono API + BullMQ Worker hosting",            "Starter",          "Active",      "€15",   LIGHT_BLUE),
    ("GitHub",              "github.com/chrisndi-syn",   "chrisndi-syn",                "Source code — SyndicSage repo (go private!)", "Free → Team",      "Public ⚠️",   "€0 → €4", WHITE),
    ("Google Workspace",    "workspace.google.com",      "hello@syndicsage.com",        "Business email + Google Sign In OAuth",       "Business Starter", "Active",      "€6",    LIGHT_BLUE),
    ("Anthropic",           "console.anthropic.com",     "chris account",               "Claude API — AI Sage features",               "Pay-as-you-go",    "Ready",       "€20+",  WHITE),
    ("Namecheap",           "namecheap.com",             "chris account",               "syndicsage.com domain + DNS",                 "Domain",           "Active",      "€1/mo", LIGHT_BLUE),
    ("Resend",              "resend.com",                "chrisndi-syn account",        "Transactional email (invites, reminders)",     "Free",             "Set up",      "€0",    WHITE),
    ("Upstash",             "upstash.com",               "chrisndi-syn account",        "Redis — rate limiting + BullMQ queue",        "Free",             "Set up",      "€0",    LIGHT_BLUE),
    ("Sentry",              "sentry.io",                 "chrisndi-syn account",        "Error tracking + performance monitoring",     "Free",             "Set up",      "€0",    WHITE),
    ("Apple Developer",     "developer.apple.com",       "chris Apple ID",              "App Store publishing + Apple Sign In",        "Standard ($99/yr)","Phase 7",     "€8/mo", LIGHT_BLUE),
    ("Google Play Console", "play.google.com/console",   "chris Google account",        "Android app publishing",                      "One-time €25",     "Phase 7",     "€0",    WHITE),
    ("EUIPO",               "euipo.europa.eu",           "chris account",               "SyndicSage trademark filing (Class 42+35)",   "€900 one-time",    "Pre-launch",  "€0",    LIGHT_AMBER),
    ("iubenda",             "iubenda.com",               "chris account",               "GDPR privacy policy + cookie policy",         "€27/yr",           "Pre-launch",  "€2/mo", WHITE),
    ("PostHog",             "posthog.com",               "chrisndi-syn account",        "Product analytics (optional)",                "Free",             "Optional",    "€0",    LIGHT_BLUE),
]

for i, row_data in enumerate(accounts):
    r = 3 + i
    service, url, account, purpose, tier, status, cost, bg = row_data
    values = [service, url, account, purpose, tier, status, cost]
    for col, val in enumerate(values, 1):
        c = ws4.cell(row=r, column=col, value=val)
        c.fill = fill(bg)
        c.border = border()
        c.alignment = align("left", wrap=True)
        c.font = font(size=10)
    # Highlight warning
    if "⚠️" in status:
        ws4.cell(row=r, column=6).font = font(bold=True, color=RED, size=10)
    ws4.row_dimensions[r].height = 30

for col, w in [(1,22),(2,30),(3,28),(4,40),(5,22),(6,14),(7,14)]:
    set_col_width(ws4, col, w)

# ── SHEET 5: PHASE ROLLOUT ────────────────────────────────────
ws5 = wb.create_sheet("Phase Rollout")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:H1")
c = ws5["A1"]
c.value = "Cost Rollout by Phase"
c.fill = fill(NAVY); c.font = Font(bold=True, color=WHITE, size=14)
c.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 32

header_row(ws5, 2, ["Phase", "Name", "Services Activated", "New Monthly Cost", "Cumulative Monthly", "One-Time This Phase", "Cumulative One-Time", "Notes"])

rollout = [
    ("Phase 0",   "Architecture Foundation", "Supabase Pro, Railway API, Railway Worker, Upstash Redis (free)", 40,  40,   0,    0,    "Core infrastructure spun up. No users yet.", LIGHT_BLUE),
    ("Phase 1",   "Shell",                   "GitHub Team (private repo), Google Workspace",                     21,  61,   0,    0,    "Login page live. Repo goes private.", WHITE),
    ("Phase 2",   "Core Data",               "No new services",                                                   0,   61,   0,    0,    "Buildings, owners, charges, settings.", LIGHT_BLUE),
    ("Phase 2.5", "Audit Foundation",         "No new services",                                                   0,   61,   0,    0,    "Audit log + GDPR table created.", WHITE),
    ("Phase 3",   "Communication",           "Resend (free), Sentry (free)",                                      0,   61,   0,    0,    "Documents, inbox, timeline, tickets.", LIGHT_BLUE),
    ("Phase 4",   "Intelligence",            "Anthropic API",                                                     20,  81,   0,    0,    "AI Sage, extraction, RAG. Cost scales with usage.", WHITE),
    ("Phase 5",   "Governance",              "No new services",                                                    0,   81,   0,    0,    "Roadmap, voting, meetings, reporting.", LIGHT_BLUE),
    ("Phase 6",   "Portal",                  "Resend Scale (more emails)",                                        20,  101,  0,    0,    "Resident portal live. Email volume increases.", WHITE),
    ("Phase 7",   "Mobile",                  "Apple Developer ($99/yr), Google Play (€25 once)",                  8,   109,  124,  124,  "iOS + Android. Apple Dev €99/yr = €8/mo equiv.", LIGHT_BLUE),
    ("Phase 8",   "Launch",                  "Vercel Pro, iubenda, trademark filed",                              29,  138,  927,  1051, "Vercel Pro €20 + iubenda €2 + GitHub Team already counted.", WHITE),
]

for i, row_data in enumerate(rollout):
    r = 3 + i
    phase, name, services, new_mo, cum_mo, new_once, cum_once, notes, bg = row_data
    values = [phase, name, services, new_mo, cum_mo, new_once, cum_once, notes]
    for col, val in enumerate(values, 1):
        c = ws5.cell(row=r, column=col, value=val)
        c.fill = fill(bg)
        c.border = border()
        c.alignment = align("left", wrap=True)
        c.font = font(size=10)
    for col in [4, 5, 6, 7]:
        ws5.cell(row=r, column=col).number_format = u'€#,##0.00'
        ws5.cell(row=r, column=col).alignment = align("right")
    ws5.row_dimensions[r].height = 36

for col, w in [(1,12),(2,24),(3,46),(4,18),(5,20),(6,20),(7,20),(8,40)]:
    set_col_width(ws5, col, w)

# ── SHEET 6: BREAK-EVEN ───────────────────────────────────────
ws6 = wb.create_sheet("Break-Even")
ws6.sheet_view.showGridLines = False

ws6.merge_cells("A1:E1")
c = ws6["A1"]
c.value = "Break-Even & Revenue Projection"
c.fill = fill(NAVY); c.font = Font(bold=True, color=WHITE, size=14)
c.alignment = Alignment(horizontal="center", vertical="center")
ws6.row_dimensions[1].height = 32

# Pricing assumptions
ws6.cell(row=2, column=1, value="Pricing Assumptions").font = font(bold=True, color=NAVY, size=12)
ws6.row_dimensions[2].height = 22

assumptions = [
    ("Starter plan price",      49,  "€/month per customer"),
    ("Pro plan price",          99,  "€/month per customer"),
    ("Enterprise plan price",   199, "€/month per customer"),
    ("Monthly costs (MVP)",     61,  "€/month (Phase 0-3 baseline)"),
    ("Monthly costs (Launch)",  138, "€/month (Phase 8 full)"),
]
header_row(ws6, 3, ["Parameter", "Value", "Unit", "", ""])
for i, (label, val, unit) in enumerate(assumptions):
    r = 4 + i
    bg = LIGHT_BLUE if i % 2 == 0 else WHITE
    for col, v in enumerate([label, val, unit], 1):
        c = ws6.cell(row=r, column=col, value=v)
        c.fill = fill(bg); c.border = border(); c.font = font(size=10)
        c.alignment = align("left")
    ws6.cell(row=r, column=2).number_format = u'€#,##0.00'
    ws6.cell(row=r, column=2).alignment = align("right")
    ws6.row_dimensions[r].height = 22

ws6.row_dimensions[9].height = 16

# Break-even table
ws6.cell(row=10, column=1, value="Break-Even Analysis").font = font(bold=True, color=NAVY, size=12)
ws6.row_dimensions[10].height = 22

ws6.cell(row=11, column=1, value="Break-even (MVP costs, Starter plan)").font = font(bold=True, color=NAVY)
c = ws6.cell(row=11, column=2, value="=CEILING(B8/B4,1)")
c.number_format = '0 "customers"'
c.font = font(bold=True, color=GREEN, size=13)
c.fill = fill(LIGHT_GREEN); c.border = border(); c.alignment = align("right")
ws6.row_dimensions[11].height = 26

ws6.cell(row=12, column=1, value="Break-even (Launch costs, Starter plan)").font = font(bold=True, color=NAVY)
c = ws6.cell(row=12, column=2, value="=CEILING(B9/B4,1)")
c.number_format = '0 "customers"'
c.font = font(bold=True, color=AMBER, size=13)
c.fill = fill(LIGHT_AMBER); c.border = border(); c.alignment = align("right")
ws6.row_dimensions[12].height = 26

ws6.row_dimensions[13].height = 16

# Projection table
ws6.cell(row=14, column=1, value="Revenue Projection").font = font(bold=True, color=NAVY, size=12)
ws6.row_dimensions[14].height = 22
header_row(ws6, 15, ["Customers", "Starter (€49)", "Pro (€99)", "Mixed (50/50)", "Profit (Mixed - Launch costs)"])

for i, n in enumerate([1, 2, 3, 5, 10, 20, 50, 100]):
    r = 16 + i
    bg = LIGHT_GREEN if n >= 3 else LIGHT_RED
    ws6.cell(row=r, column=1, value=n).font = font(size=11, bold=(n >= 3))
    eur(ws6, r, 2, formula=f"=A{r}*B4", bg=bg)
    eur(ws6, r, 3, formula=f"=A{r}*B5", bg=bg)
    eur(ws6, r, 4, formula=f"=A{r}*(B4+B5)/2", bg=bg)
    c = ws6.cell(row=r, column=5, value=f"=D{r}-B9")
    c.number_format = u'€#,##0.00'
    profit_bg = LIGHT_GREEN if n >= 3 else LIGHT_RED
    c.fill = fill(profit_bg)
    c.font = font(bold=(n >= 3), color=GREEN if n >= 3 else RED, size=11)
    c.alignment = align("right"); c.border = border()
    ws6.row_dimensions[r].height = 22

# Summary note
note_row = 16 + 8 + 1
ws6.merge_cells(f"A{note_row}:E{note_row}")
c = ws6.cell(row=note_row, column=1,
             value="Note: Break-even is at 3 customers on Starter plan (MVP phase). At launch costs, 3 customers still covers it. First €49 customer = profitable at MVP stage.")
c.font = font(italic=True, color=TEXT_GRAY, size=10)
c.fill = fill(LIGHT_GREEN)
c.alignment = Alignment(horizontal="left", wrap_text=True)
ws6.row_dimensions[note_row].height = 36

for col, w in [(1,40),(2,16),(3,16),(4,18),(5,28)]:
    set_col_width(ws6, col, w)

# ── SHEET 7: DEFERRED DECISIONS ───────────────────────────────
ws7 = wb.create_sheet("Deferred Costs")
ws7.sheet_view.showGridLines = False

ws7.merge_cells("A1:E1")
c = ws7["A1"]
c.value = "Deferred & Future Costs (not in MVP budget)"
c.fill = fill(NAVY); c.font = Font(bold=True, color=WHITE, size=14)
c.alignment = Alignment(horizontal="center", vertical="center")
ws7.row_dimensions[1].height = 32

header_row(ws7, 2, ["Item", "Estimated Cost", "Frequency", "Trigger", "Notes"])

deferred = [
    ("Self-hosted Supabase (EU VPS)",    "€20-50/mo",    "Monthly",   "Enterprise client requires on-premise",          "Supabase is open source — self-host entire stack on Hetzner/OVH EU VPS.", LIGHT_AMBER),
    ("MinIO (document storage)",         "€10-30/mo",    "Monthly",   "On-premise contract or DPA requirement",          "S3-compatible. Same path structure. Zero code changes.", WHITE),
    ("itsme® authentication",            "€0.10-0.30/tx","Per login",  "Enterprise/government customer request",         "OIDC — plugs into Supabase as custom OAuth. Contact itsme.be/business.", LIGHT_AMBER),
    ("External pentest",                 "€1,500-3,000", "One-time",  "First enterprise contract or €50k ARR",          "OWASP Top 10 + tenant isolation + file upload attack vectors.", WHITE),
    ("LaunchDarkly (feature flags)",     "€0-350/mo",    "Monthly",   "Team grows beyond 3 devs",                       "Current plan: feature_flags Supabase table (free). LaunchDarkly when ops complexity grows.", LIGHT_AMBER),
    ("Read replica (Supabase)",          "~€25/mo extra","Monthly",   "DB query latency > 200ms on list endpoints",      "Supabase one-click. Route reads to replica, writes to primary.", WHITE),
    ("Kubernetes / EKS",                 "€200+/mo",     "Monthly",   "Railway costs > €500/mo or GPU workloads needed","Hono is container-native — zero code changes to migrate.", LIGHT_AMBER),
    ("SOC 2 / ISO 27001 audit",          "€5,000-20,000","One-time",  "Enterprise procurement checklist",               "Architecture already designed to pass. AWS accounts needed for SOC 2.", WHITE),
    ("Field-level encryption (KMS)",     "€50-100/mo",   "Monthly",   "Enterprise DPA requirement for IBAN/national IDs","packages/crypto + KMS-backed envelope encryption. Additive change.", LIGHT_AMBER),
    ("Apple Developer (mobile)",         "€99/yr",       "Annual",    "Phase 7 — mobile app",                           "Required for App Store. Already in One-Time budget.", WHITE),
    ("Expo EAS Production build",        "$99/mo",       "Monthly",   "Mobile app CI/CD at scale",                      "Free tier: 30 builds/mo. Upgrade when release cadence increases.", LIGHT_AMBER),
    ("Multi-region (Vercel + Railway)",  "€50-100/mo",   "Monthly",   "Customers outside Belgium require low latency",   "Vercel edge already global. Railway: add EU + US regions.", WHITE),
]

for i, row_data in enumerate(deferred):
    r = 3 + i
    item, cost, freq, trigger, notes, bg = row_data
    for col, val in enumerate([item, cost, freq, trigger, notes], 1):
        c = ws7.cell(row=r, column=col, value=val)
        c.fill = fill(bg); c.border = border()
        c.alignment = align("left", wrap=True)
        c.font = font(size=10)
    ws7.row_dimensions[r].height = 36

for col, w in [(1,32),(2,18),(3,14),(4,36),(5,52)]:
    set_col_width(ws7, col, w)

# ── Save ──────────────────────────────────────────────────────
output_path = "/Users/chris/syndicsage/docs/v5-costs.xlsx"
wb.save(output_path)
print(f"Saved → {output_path}")
